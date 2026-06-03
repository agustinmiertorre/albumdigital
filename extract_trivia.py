import openpyxl, json

wb = openpyxl.load_workbook('preguntas_trivia.xlsx')

all_data = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append([str(c).strip() if c is not None else None for c in row])

    # Find the blank row separator between ES and EN sections
    blank_row_idx = None
    for i, row in enumerate(rows):
        if all(c is None for c in row) and i > 10:
            blank_row_idx = i
            break

    def parse_section(row_list):
        questions = []
        for row in row_list:
            if not row[0]:
                continue
            r0 = row[0]
            # Skip header rows
            if 'Pregunta' in r0 or 'Question' in r0 or 'Respuesta' in r0 or 'Correct' in r0 or 'Incorrect' in r0:
                continue
            q_text = r0
            correct = str(row[1]).strip() if row[1] else ''
            wrong = [str(row[j]).strip() for j in range(2, 5) if row[j] is not None]
            if not correct or correct == 'None':
                continue
            # Skip Excel datetime artifacts
            if '2026-0' in correct:
                continue
            # Fix fractions stored as dates
            if '1/2' in q_text and '2026' not in str(row[1]):
                pass
            questions.append({'q': q_text, 'ok': correct, 'ops': wrong})
        return questions

    if blank_row_idx is not None:
        es_rows = rows[1:blank_row_idx]  # skip header row 0
        en_rows = rows[blank_row_idx+1:]
    else:
        # No blank separator - split by half
        mid = (len(rows) - 1) // 2
        es_rows = rows[1:mid+1]
        en_rows = rows[mid+1:]

    es_questions = parse_section(es_rows)
    en_questions = parse_section(en_rows)

    grade = '1'
    for ch in sheet_name:
        if ch.isdigit():
            grade = ch
            break

    all_data[grade] = {'es': es_questions, 'en': en_questions}
    print(f"Hoja {sheet_name} -> grado {grade}: {len(es_questions)} ES, {len(en_questions)} EN")
    if es_questions:
        print(f"  ES[0]: {es_questions[0]['q'][:60]}")
    if en_questions:
        print(f"  EN[0]: {en_questions[0]['q'][:60]}")

with open('trivia_data.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)
print("\nGuardado en trivia_data.json")
