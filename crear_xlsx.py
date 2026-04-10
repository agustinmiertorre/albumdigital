import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# === HOJA 1: TRIVIA ===
ws = wb.active
ws.title = 'Trivia'

headers = ['ID','Pregunta','Respuesta Correcta','Opcion 2','Opcion 3','Opcion 4','Nivel Min (1-6)','Nivel Max (1-6)','Categoria']
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = PatternFill('solid', fgColor='003366')
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

rows = [
  # NIVEL 1-2 (6-7 años)
  ['T001','De que color es el pasto de una cancha de futbol?','Verde','Azul','Rojo','Amarillo',1,2,'cultura_general'],
  ['T002','Cuantos jugadores hay en un equipo de futbol?','11','5','22','10',1,2,'mundial'],
  ['T003','Que forma tiene la pelota de futbol?','Redonda','Cuadrada','Triangular','Ovalada',1,2,'mundial'],
  ['T004','Cuanto es 2 + 3?','5','4','6','7',1,2,'escolar'],
  ['T005','De que colores es la camiseta de Argentina?','Azul y blanca','Roja y negra','Verde y amarilla','Naranja y azul',1,2,'mundial'],
  ['T006','Cuantas letras tiene la palabra GATO?','4','3','5','2',1,2,'escolar'],
  ['T007','Que obtienes mezclando azul y amarillo?','Verde','Naranja','Violeta','Marron',1,2,'escolar'],
  ['T008','En que continente esta Argentina?','America','Europa','Asia','Africa',1,2,'mundial'],
  ['T009','Cuantos dedos tiene una mano?','5','4','6','10',1,2,'escolar'],
  ['T010','Cual es el animal mas grande del mundo?','Ballena azul','Elefante','Jirafa','Hipopotamo',1,2,'cultura_general'],
  # NIVEL 3-4 (8-10 años)
  ['T011','Cuantos paises participan en el Mundial 2026?','48','32','24','64',3,4,'mundial'],
  ['T012','En que pais nacio Lionel Messi?','Argentina','Brasil','Uruguay','Espana',3,4,'cultura_general'],
  ['T013','Quien gano el Mundial de futbol de 2022?','Argentina','Francia','Brasil','Alemania',3,4,'mundial'],
  ['T014','Cual es la capital de Brasil?','Brasilia','Rio de Janeiro','Sao Paulo','Buenos Aires',3,4,'mundial'],
  ['T015','En que continente esta Marruecos?','Africa','Asia','America','Europa',3,4,'mundial'],
  ['T016','Cuanto es 7 x 8?','56','48','64','54',3,4,'escolar'],
  ['T017','Cuantos dias tiene un ano no bisiesto?','365','360','366','364',3,4,'escolar'],
  ['T018','Cuanto es 100 dividido 4?','25','20','30','40',3,4,'escolar'],
  ['T019','Cual es el idioma oficial de Francia?','Frances','Ingles','Espanol','Aleman',3,4,'mundial'],
  ['T020','Cuantos lados tiene un hexagono?','6','5','7','8',3,4,'escolar'],
  ['T021','Que planeta es el mas cercano al Sol?','Mercurio','Venus','La Tierra','Marte',3,4,'escolar'],
  ['T022','En que continente esta Japon?','Asia','America','Europa','Oceania',3,4,'mundial'],
  ['T023','Cuantos colores tiene el arcoiris?','7','5','6','8',3,4,'cultura_general'],
  # NIVEL 5-6 (11-13 años)
  ['T024','En que ano se jugo el primer Mundial de futbol?','1930','1920','1950','1910',5,6,'mundial'],
  ['T025','Que pais gano mas Mundiales de futbol?','Brasil con 5','Alemania con 4','Italia con 4','Argentina con 3',5,6,'mundial'],
  ['T026','Cuantos titulos mundiales tiene Alemania?','4','3','5','2',5,6,'mundial'],
  ['T027','Cual es la moneda oficial de Japon?','Yen','Yuan','Won','Dolar',5,6,'mundial'],
  ['T028','Cuanto es el 15% de 200?','30','25','20','40',5,6,'escolar'],
  ['T029','Cual es el perimetro de un cuadrado de lado 5 cm?','20 cm','25 cm','10 cm','15 cm',5,6,'escolar'],
  ['T030','Quien escribio Don Quijote de la Mancha?','Miguel de Cervantes','Shakespeare','Borges','Garcia Marquez',5,6,'escolar'],
  ['T031','Cuantos paises tiene America del Sur?','12','10','14','8',5,6,'cultura_general'],
  ['T032','En que ano Argentina gano su segundo Mundial?','1986','1978','1990','1982',5,6,'mundial'],
  ['T033','Cual es la formula del area de un triangulo?','Base x Altura / 2','Base x Altura','Base + Altura','2 x (Base + Altura)',5,6,'escolar'],
  ['T034','Que gas necesitan los seres vivos para respirar?','Oxigeno','Nitrogeno','Dioxido de carbono','Hidrogeno',5,6,'escolar'],
  ['T035','Cuantas selecciones clasifica Europa al Mundial 2026?','16','12','10','8',5,6,'mundial'],
]

for row_data in rows:
    ws.append(row_data)

cat_colors = {'cultura_general':'E8F4FD','mundial':'E8F8E8','escolar':'FFF8E1'}
for row in ws.iter_rows(min_row=2):
    cat = row[8].value if len(row) > 8 and row[8].value else ''
    color = cat_colors.get(cat, 'FFFFFF')
    for cell in row:
        cell.fill = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(wrap_text=True, vertical='center')

widths = [8,55,30,20,20,20,14,14,18]
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30

# === HOJA 2: SABIAS QUE ===
ws2 = wb.create_sheet('Sabias Que')
ws2.append(['ID','Dato curioso','Nivel Min','Nivel Max','Categoria'])
for col in range(1,6):
    c = ws2.cell(1,col)
    c.fill = PatternFill('solid', fgColor='5B2D8E')
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.alignment = Alignment(horizontal='center')

sabias = [
  ['S001','Brasil es el unico pais que jugo todos los Mundiales de futbol de la historia.',1,6,'mundial'],
  ['S002','La Copa del Mundo pesa 6.175 kg y esta hecha de oro de 18 quilates.',1,6,'mundial'],
  ['S003','El Mundial 2026 sera el primero con 48 selecciones en lugar de 32.',1,6,'mundial'],
  ['S004','El gol mas rapido en un Mundial lo marco Hakan Sukur en solo 11 segundos (Corea 2002).',3,6,'mundial'],
  ['S005','Argentina gano su tercer Mundial en 2022 en Qatar, venciendo a Francia en la final.',1,6,'mundial'],
  ['S006','Japon tiene mas de 6.800 islas en su territorio.',3,6,'mundial'],
  ['S007','Marruecos fue el primer pais africano en llegar a las semifinales de un Mundial (Qatar 2022).',3,6,'mundial'],
  ['S008','El estadio mas grande del Mundial 2026 tiene capacidad para mas de 100.000 personas.',1,6,'mundial'],
  ['S009','El album de figuritas del Mundial fue creado por Panini en Italia en 1970.',1,6,'mundial'],
  ['S010','En Argentina se venden cientos de millones de figuritas del Mundial cada edicion.',1,6,'mundial'],
  ['S011','Portugal tiene 10 millones de habitantes, pero mas de 250 millones hablan portugues en el mundo.',5,6,'mundial'],
  ['S012','Un partido de futbol tiene 90 minutos divididos en dos tiempos de 45 minutos.',1,2,'mundial'],
  ['S013','Mexico organizara el Mundial 2026 junto con EE.UU. y Canada.',1,6,'mundial'],
  ['S014','Australia clasifico a 5 Mundiales consecutivos entre 2006 y 2022.',3,6,'mundial'],
  ['S015','El idioma coreano (hangul) fue inventado por el Rey Sejong en el siglo XV.',5,6,'mundial'],
  ['S016','Francia es el pais con mas jugadores nacidos en el exterior en su seleccion.',5,6,'mundial'],
  ['S017','Senegal gano su primera Copa de Africa en 2021 con Sadio Mane como figura.',3,6,'mundial'],
  ['S018','La bandera de Uruguay tiene el mismo celeste que Argentina pero con franjas horizontales.',3,6,'mundial'],
]

for row_data in sabias:
    ws2.append(row_data)
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='center')
ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 80
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 15

# === HOJA 3: INSTRUCCIONES ===
ws3 = wb.create_sheet('Instrucciones')
instr = [
    ['INSTRUCCIONES PARA EDITAR LAS PREGUNTAS'],
    [''],
    ['HOJA TRIVIA:'],
    ['  ID: identificador unico (no repetir). Ej: T036'],
    ['  Pregunta: texto de la pregunta'],
    ['  Respuesta Correcta: la respuesta que da el punto'],
    ['  Opcion 2/3/4: respuestas incorrectas'],
    ['  Nivel Min/Max: rango de grados (1=primer grado, 6=sexto grado)'],
    ['  Categoria: cultura_general / mundial / escolar'],
    [''],
    ['HOJA SABIAS QUE:'],
    ['  Los datos aparecen en el encabezado de la app rotando cada 10 segundos'],
    [''],
    ['NIVELES RECOMENDADOS:'],
    ['  1 y 2 grado (6-7 anos): Nivel 1-2 - preguntas muy simples'],
    ['  3 y 4 grado (8-10 anos): Nivel 3-4 - conocimiento general'],
    ['  5 y 6 grado (11-13 anos): Nivel 5-6 - preguntas mas complejas'],
    [''],
    ['Para aplicar cambios en la app, contactar al Departamento de Tecnologia.'],
]
for r in instr:
    ws3.append(r)
ws3.column_dimensions['A'].width = 80
ws3['A1'].font = Font(bold=True, size=13, color='003366')

wb.save(r'C:/Users/Aula Tech/Desktop/Proyecto/Album de figuritas/preguntas_trivia.xlsx')
print('OK - preguntas_trivia.xlsx creado')
